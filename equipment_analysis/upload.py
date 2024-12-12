import os
import subprocess  # For running shell commands
import pandas as pd
import msal
from office365.sharepoint.client_context import ClientContext
from io import BytesIO
import openpyxl  # For manipulating Excel files
import project_config  # Your config file with site_url, client_id, client_secret, tenant_id

def sync_and_update_excel():
    # Configuration from project_config
    site_url = project_config.site_url
    client_id = project_config.client_id
    client_secret = project_config.client_secret
    tenant_id = project_config.tenant_id
    target_file_name = project_config.target_file_name
    sheet_name = project_config.sheet_name
    local_new_data_path = 'aggregated_data.xlsx'
    onedrive_folder_path = project_config.onedrive_path

    target_file_path = os.path.join(onedrive_folder_path, target_file_name)

    scope = ["https://graph.microsoft.com/.default"]

    def sync_onedrive():
        print("Syncing OneDrive...")
        try:
            result = subprocess.run(["onedrive", "--synchronize"], capture_output=True, text=True)
            if result.returncode == 0:
                print("OneDrive synced successfully.")
            else:
                print(f"OneDrive sync failed: {result.stderr}")
        except Exception as e:
            print(f"Error syncing OneDrive: {e}")

    print("Authenticating to SharePoint...")
    app = msal.ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret,
    )

    print("Acquiring access token...")
    result = app.acquire_token_for_client(scopes=scope)

    if "access_token" in result:
        print("Access token acquired successfully.")
        
        ctx = ClientContext(site_url).with_access_token(result["access_token"])

        sync_onedrive()

        def load_or_create_sheet(workbook, sheet_name):
            if sheet_name in workbook.sheetnames:
                print(f"Sheet '{sheet_name}' exists, loading data...")
                return workbook[sheet_name]
            else:
                print(f"Sheet '{sheet_name}' does not exist, creating new sheet...")
                return workbook.create_sheet(sheet_name)

        def extract_existing_dates(sheet):
            try:
                return {
                    pd.to_datetime(sheet.cell(row=i, column=1).value, errors='coerce').date()
                    for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value
                }
            except Exception as e:
                print(f"Error extracting existing dates: {e}")
                return set()

        if os.path.exists(target_file_path):
            print(f"Opening the existing Excel file from '{target_file_path}'...")
            workbook = openpyxl.load_workbook(target_file_path)
            existing_sheet = load_or_create_sheet(workbook, sheet_name)
        else:
            print(f"File '{target_file_name}' does not exist locally. Downloading from SharePoint...")
            folder_url = project_config.folder_url
            operations_folder = ctx.web.get_folder_by_server_relative_url(folder_url)
            ctx.load(operations_folder)
            ctx.execute_query()

            files_collection = operations_folder.files
            ctx.load(files_collection)
            ctx.execute_query()

            target_file_found = False
            for file_item in files_collection:
                if file_item.properties['Name'] == target_file_name:
                    target_file_found = True
                    target_file = file_item
                    break

            if target_file_found:
                print(f"File '{target_file_name}' found in SharePoint. Downloading for local processing...")
                existing_file_stream = BytesIO()
                target_file.download(existing_file_stream).execute_query()
                existing_file_stream.seek(0)
                workbook = openpyxl.load_workbook(existing_file_stream)
                existing_sheet = load_or_create_sheet(workbook, sheet_name)
            else:
                print(f"Target file '{target_file_name}' not found in the SharePoint folder.")
                return  # Use return instead of exit()

        new_data_df = pd.read_excel(local_new_data_path)
        print(f"New data loaded from '{local_new_data_path}'.")

        if existing_sheet.max_row == 1:
            print("Inserting column headers...")
            existing_sheet.append(list(new_data_df.columns))

        existing_dates = extract_existing_dates(existing_sheet)

        # Ensure a blank row is added if there is data in the sheet
        if existing_sheet.max_row > 1:
            print("Inserting a blank row before appending new data...")
            existing_sheet.append([])

        new_data_list = new_data_df.values.tolist()
        next_row = existing_sheet.max_row + 1
        print(f"Appending new data starting from row {next_row}...")

        for row in new_data_list:
            try:
                date_value = pd.to_datetime(row[0], errors='coerce').date()
                if row[0].strip().lower() == "total" or date_value not in existing_dates:
                    existing_sheet.append(row)
                    print(f"Added new data for date or total: {row}")
                else:
                    print(f"Skipping existing date: {date_value}")
            except Exception as e:
                print(f"Error processing row {row}: {e}")

        # Save changes to the workbook
        workbook.save(target_file_path)
        print(f"Updated data saved to '{target_file_path}'.")
    else:
        print("Error acquiring token:", result.get("error"), result.get("error_description"))

if __name__ == "__main__":
    sync_and_update_excel()
